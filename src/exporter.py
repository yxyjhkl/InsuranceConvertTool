"""
导出模块
支持Excel、图片、A4纸张格式导出
"""

import os
import sys
import io
import logging
from typing import List, Dict, Optional, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


def find_font(size: int = 12, bold: bool = False) -> Optional[str]:
    """查找可用的中文字体（桌面端用 Windows/Linux/macOS，Android 用默认）"""
    candidates = []
    if sys.platform == 'win32':
        font_dir = os.environ.get('WINDIR', 'C:\\Windows') + '\\Fonts'
        if bold:
            candidates = [
                os.path.join(font_dir, 'msyhbd.ttc'),
                os.path.join(font_dir, 'simhei.ttf'),
                os.path.join(font_dir, 'simsun.ttc'),
            ]
        else:
            candidates = [
                os.path.join(font_dir, 'msyh.ttc'),
                os.path.join(font_dir, 'simhei.ttf'),
                os.path.join(font_dir, 'simsun.ttc'),
            ]
    elif sys.platform == 'darwin':
        candidates = [
            '/System/Library/Fonts/PingFang.ttc',
            '/System/Library/Fonts/STHeiti Light.ttc',
        ]
    elif hasattr(sys, 'getandroidapilevel'):
        # Android: 系统自带 DroidSansFallback 或 Noto 字体
        candidates = [
            '/system/fonts/DroidSansFallback.ttf',
            '/system/fonts/NotoSansCJK-Regular.ttc',
            '/system/fonts/NotoSansSC-Regular.otf',
        ]
    else:
        candidates = [
            '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        ]

    for path in candidates:
        if os.path.exists(path):
            return path
    return None


class ExcelExporter:
    """Excel导出器"""
    COLUMN_WIDTHS = [10, 8, 12, 12, 14, 12, 12, 14, 14, 14, 12, 14, 12, 12]

    def __init__(self, headers: List[str], data: List[List], customer_name: str = ''):
        self.headers = headers
        self.data = data
        self.customer_name = customer_name

    def export(self, filepath: str):
        wb = Workbook()
        ws = wb.active
        ws.title = '保险建议书'

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        cell_font = Font(name='微软雅黑', size=10)
        cell_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, (header, width) in enumerate(zip(self.headers, self.COLUMN_WIDTHS), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        for row_idx, row_data in enumerate(self.data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        if self.customer_name:
            ws.insert_rows(1)
            info_cell = ws.cell(row=1, column=1, value=f'客户: {self.customer_name}')
            info_cell.font = Font(name='微软雅黑', size=12, bold=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.headers))

        wb.save(filepath)


class ImageExporter:
    """图片导出器"""
    A4_WIDTH = 2480
    A4_HEIGHT = 3508
    HEADER_HEIGHT = 60
    ROW_HEIGHT = 45
    PADDING = 40

    def __init__(self, headers: List[str], data: List[List], customer_name: str = ''):
        self.headers = headers
        self.data = data
        self.customer_name = customer_name
        self.font_path = find_font(12)
        self.bold_font_path = find_font(14, True)

    def _get_font(self, size: int, bold: bool = False):
        try:
            if bold:
                return ImageFont.truetype(self.bold_font_path or '', size)
            return ImageFont.truetype(self.font_path or '', size)
        except Exception:
            return ImageFont.load_default()

    def _draw_table(self, draw: ImageDraw.Draw, start_x: int, start_y: int,
                    col_widths: List[int]) -> Tuple[int, int]:
        font = self._get_font(11)
        bold_font = self._get_font(12, True)
        small_font = self._get_font(10)
        P = self.PADDING

        if self.customer_name:
            draw.text((start_x + P, start_y + 10),
                     f'客户: {self.customer_name}', fill='#1F4E79', font=bold_font)
            start_y += 40

        header_y = start_y
        x = start_x
        for i, (header, width) in enumerate(zip(self.headers, col_widths)):
            draw.rectangle([x, header_y, x + width, header_y + self.HEADER_HEIGHT],
                          fill='#2F5496', outline='#1F4E79', width=1)
            text = header[:6]
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_x = x + (width - text_width) // 2
            draw.text((text_x, header_y + 18), text, fill='white', font=font)
            x += width

        start_y += self.HEADER_HEIGHT

        for row_idx, row in enumerate(self.data):
            x = start_x
            row_y = start_y
            bg_color = '#F2F2F2' if row_idx % 2 == 0 else '#FFFFFF'

            for col_idx, (value, width) in enumerate(zip(row, col_widths)):
                draw.rectangle([x, row_y, x + width, row_y + self.ROW_HEIGHT],
                              fill=bg_color, outline='#CCCCCC', width=1)
                text = str(value) if value else ''
                if len(text) > 10:
                    text = text[:9] + '..'
                bbox = draw.textbbox((0, 0), text, font=small_font)
                text_width = bbox[2] - bbox[0]
                text_x = x + (width - text_width) // 2
                draw.text((text_x, row_y + 14), text, fill='#333333', font=small_font)
                x += width
            start_y += self.ROW_HEIGHT

        return start_x + sum(col_widths), start_y

    def export_image(self, filepath: str, width: int = None, height: int = None):
        col_widths = [200, 150, 180, 180, 200, 180, 180, 200, 200, 200, 180, 200, 180, 180]
        table_width = sum(col_widths) + self.PADDING * 2
        table_height = self.ROW_HEIGHT * (len(self.data) + 1) + 100

        if self.customer_name:
            table_height += 40

        img_height = max(height or 0, table_height + self.PADDING * 2)
        img_width = max(width or 0, table_width)

        img = Image.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(img)

        title_font = self._get_font(20, True)
        title = '保险建议书数据'
        bbox = draw.textbbox((0, 0), title, font=title_font)
        title_width = bbox[2] - bbox[0]
        draw.text(((img_width - title_width) // 2, 20), title, fill='#1F4E79', font=title_font)

        self._draw_table(draw, self.PADDING, 80, col_widths)
        img.save(filepath, 'PNG', quality=95)

    def export_a4(self, filepath: str):
        self.export_image(filepath, self.A4_WIDTH, self.A4_HEIGHT)

    def export_screenshot(self, filepath: str):
        col_widths = [200, 150, 180, 180, 200, 180, 180, 200, 200, 200, 180, 200, 180, 180]
        table_width = sum(col_widths) + self.PADDING * 2
        max_rows = min(len(self.data), 50)
        img_height = self.ROW_HEIGHT * (max_rows + 2) + 150
        self.export_image(filepath, table_width, img_height)


class PDFToExcelConverter:
    """PDF转Excel转换器"""

    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path

    def convert(self, excel_path: str) -> bool:
        try:
            import pdfplumber
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = 'PDF数据'

            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_font = Font(name='微软雅黑', size=10)
            cell_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            current_row = 1
            max_col_widths = {}

            with pdfplumber.open(self.pdf_path) as pdf:
                total_pages = len(pdf.pages)

                for page_num, page in enumerate(pdf.pages, 1):
                    page_title = f'--- 第 {page_num} / {total_pages} 页 ---'
                    title_cell = ws.cell(row=current_row, column=1, value=page_title)
                    title_cell.font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
                    title_cell.alignment = Alignment(horizontal='center')
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row, end_column=15)
                    current_row += 1

                    tables = page.extract_tables() or []

                    if not tables:
                        tables = page.extract_tables(table_settings={
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines",
                        }) or []

                    if not tables:
                        text = page.extract_text()
                        if text:
                            lines = [l.strip() for l in text.split('\n') if l.strip()]
                            if lines:
                                table_data = []
                                for line in lines:
                                    parts = line.split() if line.split() else [line]
                                    if any(p.isdigit() or p in ['.', '-', '--'] for p in parts):
                                        table_data.append(parts)
                                if table_data:
                                    tables = [table_data]

                    extracted_tables = False

                    if tables:
                        for table in tables:
                            if not table or len(table) == 0:
                                continue
                            valid_rows = [row for row in table if row and any(cell for cell in row if cell)]
                            if not valid_rows:
                                continue
                            extracted_tables = True

                            start_row = current_row
                            start_col = 1

                            first_row = valid_rows[0]
                            is_header = False
                            if first_row:
                                first_cell = str(first_row[0]).strip() if first_row[0] else ''
                                if not first_cell.isdigit() and first_cell not in ['年度', '年', '保单年度']:
                                    is_header = True

                            if is_header and valid_rows:
                                header_row = valid_rows[0]
                                for col_idx, cell in enumerate(header_row):
                                    cell_value = str(cell).strip() if cell else ''
                                    cell_obj = ws.cell(row=start_row, column=start_col + col_idx, value=cell_value)
                                    cell_obj.font = header_font
                                    cell_obj.fill = header_fill
                                    cell_obj.alignment = header_align
                                    cell_obj.border = thin_border
                                    max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), len(cell_value))
                                data_start = 1
                            else:
                                data_start = 0

                            for row_idx, row in enumerate(valid_rows[data_start:], start=1):
                                for col_idx, cell in enumerate(row):
                                    cell_value = str(cell).strip() if cell else ''
                                    if cell_value and cell_value != '--':
                                        try:
                                            num_value = cell_value.replace(',', '')
                                            if '.' in num_value:
                                                cell_value = float(num_value)
                                            else:
                                                cell_value = int(float(num_value))
                                        except (ValueError, TypeError):
                                            pass

                                    cell_obj = ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=cell_value)
                                    cell_obj.font = cell_font
                                    cell_obj.alignment = cell_align
                                    cell_obj.border = thin_border
                                    str_val = str(cell_value) if cell_value else ''
                                    max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), len(str_val))

                            current_row = start_row + len(valid_rows) + 2

                    if not extracted_tables:
                        text = page.extract_text()
                        if text:
                            lines = [line.strip() for line in text.split('\n') if line.strip()]
                            if lines:
                                text_label = ws.cell(row=current_row, column=1, value='文本内容:')
                                text_label.font = Font(name='微软雅黑', size=10, bold=True, color='666666')
                                current_row += 1
                                for line in lines[:100]:
                                    parts = line.split()
                                    if len(parts) > 1:
                                        for col_idx, part in enumerate(parts[:15]):
                                            cell_obj = ws.cell(row=current_row, column=col_idx + 1, value=part)
                                            cell_obj.font = cell_font
                                            cell_obj.alignment = cell_align
                                            cell_obj.border = thin_border
                                            max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), len(part))
                                    else:
                                        cell_obj = ws.cell(row=current_row, column=1, value=line)
                                        cell_obj.font = cell_font
                                        cell_obj.alignment = Alignment(wrap_text=True)
                                        ws.merge_cells(start_row=current_row, start_column=1,
                                                      end_row=current_row, end_column=15)
                                    current_row += 1

                    current_row += 2

            for col_idx, width in max_col_widths.items():
                col_letter = get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter].width = min(max(width * 2, 10), 30)

            wb.save(excel_path)
            return True

        except Exception as e:
            logger.error(f"PDF转换失败: {e}", exc_info=True)
            return False


def export_table(headers: List[str], data: List[List],
                 filepath: str, customer_name: str = '',
                 format: str = 'xlsx'):
    if format == 'xlsx':
        exporter = ExcelExporter(headers, data, customer_name)
        exporter.export(filepath)
    elif format == 'png':
        exporter = ImageExporter(headers, data, customer_name)
        exporter.export_image(filepath)
    elif format == 'a4':
        exporter = ImageExporter(headers, data, customer_name)
        exporter.export_a4(filepath)
    elif format == 'screenshot':
        exporter = ImageExporter(headers, data, customer_name)
        exporter.export_screenshot(filepath)


def export_table_to_image(display_data: List[Dict], fmt: str = 'png') -> io.BytesIO:
    headers = ['保单年度', '年龄', '期交保费', '累计保费', '身故总利益',
               '主险现价', '现价增长率', '当年分红现价', '累计分红现价',
               '演示生存总利益', '演示增长率', '预期生存总利益',
               '预期增长率', '预期单利']

    col_widths = [90, 70, 120, 120, 130, 120, 110, 130, 130, 140, 110, 140, 110, 110]
    padding = 30
    row_height = 36
    header_height = 40

    table_width = sum(col_widths) + padding * 2
    table_height = row_height * (len(display_data) + 1) + header_height + padding * 2 + 80

    img = Image.new('RGB', (table_width, table_height), 'white')
    draw = ImageDraw.Draw(img)

    def get_font(size, bold=False):
        font_path = find_font(size, bold)
        if font_path:
            try:
                return ImageFont.truetype(font_path, size)
            except Exception:
                pass
        return ImageFont.load_default()

    title_font = get_font(20, True)
    header_font = get_font(11, True)
    cell_font = get_font(10)

    title = '保险建议书利益演示'
    try:
        bbox = draw.textbbox((0, 0), title, font=title_font)
        title_w = bbox[2] - bbox[0]
    except Exception:
        title_w = len(title) * 20
    draw.text(((table_width - title_w) // 2, 20), title, fill='#1F4E79', font=title_font)

    y = padding + 60
    x = padding
    for i, h in enumerate(headers):
        draw.rectangle([x, y, x + col_widths[i], y + header_height], fill='#1F4E79')
        try:
            bbox = draw.textbbox((0, 0), h, font=header_font)
            tw = bbox[2] - bbox[0]
        except Exception:
            tw = len(h) * 11
        draw.text((x + (col_widths[i] - tw) // 2, y + 10), h, fill='white', font=header_font)
        x += col_widths[i]

    y += header_height
    for row_idx, row in enumerate(display_data):
        bg = '#F2F2F2' if row_idx % 2 == 0 else 'white'
        x = padding
        for col_idx, h in enumerate(headers):
            key_map = {
                0: 'policy_year', 1: 'age', 2: 'premium', 3: 'total_premium',
                4: 'death_benefit', 5: 'cash_value', 6: 'growth_rate',
                7: 'current_dividend_cash', 8: 'accum_dividend_cash',
                9: 'demo_survival', 10: 'demo_rate', 11: 'expected_survival',
                12: 'expected_rate', 13: 'expected_simple_rate'
            }
            key = key_map.get(col_idx, '')
            val = row.get(key)

            if val is None:
                text = '--'
            elif key in ('growth_rate', 'demo_rate', 'expected_rate', 'expected_simple_rate'):
                text = f'{val:.2%}'
            elif key in ('policy_year', 'age'):
                text = str(int(val))
            else:
                text = f'{int(round(val)):,}'

            draw.rectangle([x, y, x + col_widths[col_idx], y + row_height], fill=bg)
            try:
                bbox = draw.textbbox((0, 0), text, font=cell_font)
                tw = bbox[2] - bbox[0]
            except Exception:
                tw = len(text) * 6
            draw.text((x + (col_widths[col_idx] - tw) // 2, y + 10), text, fill='#333333', font=cell_font)
            x += col_widths[col_idx]
        y += row_height

    x = padding
    for i in range(len(headers) + 1):
        if i < len(headers):
            draw.line([(x, padding + 60), (x, y)], fill='#CCCCCC', width=1)
            x += col_widths[i]
        else:
            draw.line([(padding, padding + 60), (padding, y)], fill='#CCCCCC', width=1)

    buf = io.BytesIO()
    img_format = 'PNG' if fmt.lower() == 'png' else 'JPEG'
    img.save(buf, format=img_format, quality=95)
    buf.seek(0)
    return buf